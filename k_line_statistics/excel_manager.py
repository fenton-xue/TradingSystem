import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelManager:
    def __init__(self, file_path):
        self.file_path = file_path

    def write_to_excel(self, date_str, kline_info):
        df = pd.DataFrame([[
            date_str,
            *kline_info
        ]], columns=[
            '日期', '开盘价', '最高点', '最低点', '收盘价',
            '长影线', '短影线', '实体', '总长'
        ])
        try:
            book = load_workbook(self.file_path)
            sheet = book.active
            for r in dataframe_to_rows(df, index=False, header=False):
                sheet.append(r)
            book.save(self.file_path)
        except FileNotFoundError:
            df.to_excel(self.file_path, index=False)
        except PermissionError:
            print(f"Permission denied: Unable to write to {self.file_path}. Please check if the file is open or if you have the necessary permissions.")
        except Exception as e:
            print(f"An error occurred: {e}")

    def delete_last_row(self):
        try:
            book = load_workbook(self.file_path)
            sheet = book.active
            if sheet.max_row > 1:
                sheet.delete_rows(sheet.max_row)
                book.save(self.file_path)
        except FileNotFoundError:
            print(f"File not found: {self.file_path}. Cannot delete last row.")
        except PermissionError:
            print(f"Permission denied: Unable to modify {self.file_path}. Please check if the file is open or if you have the necessary permissions.")
        except Exception as e:
            print(f"An error occurred: {e}")
