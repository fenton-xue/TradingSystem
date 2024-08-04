import datetime
import os

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from pandas import Timestamp


# 计算数据
def calculate_k_line(row):
    # 计算实体长度
    body = abs(row['close'] - row['open'])
    # 计算上影线长度
    upper_shadow = row['high'] - max(row['close'], row['open'])
    # 计算下影线长度
    lower_shadow = min(row['close'], row['open']) - row['low']
    # 定义长影线
    long_shadow = max(upper_shadow, lower_shadow)
    # 定义短影线
    short_shadow = min(upper_shadow, lower_shadow)
    # 定义总长
    long = row['high'] - row['low']
    return {
        'body': round(body, 2),
        'long_shadow': round(long_shadow, 2),
        'short_shadow': round(short_shadow, 2),
        'long': round(long, 2)
    }


def write_dict_to_existing_excel1(data_dict, sheet_name="Sheet1"):
    # 获取当前时间并格式化
    now = datetime.datetime.now()
    current_time = now.strftime("%Y%m%d")
    # 文件路径"E:\PrivateFile\关于交易\统计\K线形态统计.xlsx"
    file_path = f"E:\\PrivateFile\\关于交易\\统计\\{current_time}.xlsx"

    # 检查文件是否存在
    if os.path.exists(file_path):
        # 如果存在，加载工作簿
        workbook = load_workbook(filename=file_path)
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            # 如果工作表不存在，创建新的工作表
            workbook.active.title = sheet_name
    else:
        # 如果不存在，创建新的工作簿
        workbook = Workbook()
        # 可以在这里添加一些初始化操作，比如创建一个默认的工作表等
        workbook.active.title = sheet_name
        # 在该工作表中写入列名，分别为：日期，开盘价，最高价，最低价，收盘价，长影线，短影线，实体长，总长
        sheet = workbook.active
        sheet.append(["日期", "开盘价", "最高价", "最低价", "收盘价", "长影线", "短影线", "实体长", "总长"])

    # 打开一个工作簿
    workbook = openpyxl.load_workbook(file_path)
    # 激活默认的工作表
    sheet = workbook.active

    # 示例字典数据
    data = {
        'candle_begin_time_GMT8': Timestamp('2024-08-04 10:00:00'),
        'open': 60743.57,
        'high': 61117.63,
        'low': 60714.29,
        'close': 60886.95,
        'body': 143.38,
        'long_shadow': 230.68,
        'short_shadow': 29.28,
        'long': 403.34
    }
    # 将这些数据按照列名增加到工作表中
    # 获取现有表头并创建一个列名到列号的映射
    header_to_col = {}
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value is not None:
            header_to_col[cell_value] = col

    # 字典键与Excel列名的映射关系
    key_to_header = {
        "candle_begin_time_GMT8": "日期",
        "open": "开盘价",
        "high": "最高价",
        "low": "最低价",
        "close": "收盘价",
        "long_shadow": "长影线",
        "short_shadow": "短影线",
        "body": "实体长",
        "long": "总长"
    }

    # 将数据写入对应的列名下
    for key, values in data_dict.items():
        if key in key_to_header:
            column_name = key_to_header[key]
            if column_name in header_to_col:
                col_num = header_to_col[column_name]
                # 找到该列的第一个空单元格
                row_num = 2
                while sheet.cell(row=row_num, column=col_num).value is not None:
                    row_num += 1
                # 从第一个空单元格开始写入数据
                for value in values:
                    sheet.cell(row=row_num, column=col_num, value=value)
                    row_num += 1

    # 保存Excel文件
    workbook.save(file_path)


def write_dict_to_existing_excel(data_dict, sheet_name="Sheet1"):
    """
    将字典数据写入到指定的现有Excel文件的工作表中。如果文件或工作表不存在，则创建它们。

    :param data_dict: 要写入Excel的字典数据，键为数据列名，值为要写入的数据列表
    :param file_path: Excel文件的路径
    :param sheet_name: 要写入数据的工作表名称，默认为 "Sheet1"
    """
    # 字典键与Excel列名的映射关系
    key_to_header = {
        "candle_begin_time_GMT8": "日期",
        "open": "开盘价",
        "high": "最高价",
        "low": "最低价",
        "close": "收盘价",
        "long_shadow": "长影线",
        "short_shadow": "短影线",
        "body": "实体长",
        "long": "总长"
    }

    # 获取当前时间并格式化
    now = datetime.datetime.now()
    current_time = now.strftime("%Y%m%d")
    # 文件路径"E:\PrivateFile\关于交易\统计\K线形态统计.xlsx"
    file_path = f"E:\\PrivateFile\\关于交易\\统计\\{current_time}.xlsx"

    if not os.path.exists(file_path):
        # 如果文件不存在，则创建一个新的工作簿和工作表
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name

        # 写入列名
        for col, column_name in enumerate(key_to_header.values(), start=1):
            sheet.cell(row=1, column=col, value=column_name)
    else:
        # 加载现有的工作簿
        wb = openpyxl.load_workbook(file_path)

        # 获取指定名称的工作表，如果不存在则创建
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            # 写入列名
            for col, column_name in enumerate(key_to_header.values(), start=1):
                sheet.cell(row=1, column=col, value=column_name)

    # 获取现有表头并创建一个列名到列号的映射
    header_to_col = {}
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value is not None:
            header_to_col[cell_value] = col

    # 将数据写入对应的列名下
    for key, values in data_dict.items():
        if key in key_to_header:
            column_name = key_to_header[key]
            if column_name in header_to_col:
                col_num = header_to_col[column_name]
                # 找到该列的第一个空单元格
                row_num = 2
                while sheet.cell(row=row_num, column=col_num).value is not None:
                    row_num += 1
                # 从第一个空单元格开始写入数据
                for value in values:
                    sheet.cell(row=row_num, column=col_num, value=value)
                    row_num += 1

    # 保存工作簿
    wb.save(file_path)
    print(f"数据已成功写入到 {file_path} 的 {sheet_name} 工作表中")


def delete_last_row(file_path, sheet_name="Sheet1"):
    """
    删除Excel文件中指定工作表的最后一行内容。

    :param file_path: Excel文件的路径
    :param sheet_name: 要处理的工作表名称，默认为 "Sheet1"
    """
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在。")
        return

    try:
        # 加载现有的工作簿
        wb = openpyxl.load_workbook(file_path)

        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            print(f"工作表 {sheet_name} 不存在。")
            return

        sheet = wb[sheet_name]

        # 找到最后一行有数据的行号
        last_row = sheet.max_row
        while last_row > 1:
            if any(sheet.cell(row=last_row, column=col).value is not None for col in range(1, sheet.max_column + 1)):
                break
            last_row -= 1

        if last_row > 1:
            # 清除最后一行的数据
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=last_row, column=col).value = None
            print(f"已删除 {sheet_name} 工作表的最后一行内容。")
        else:
            print(f"{sheet_name} 工作表没有可删除的行。")

        # 保存工作簿
        wb.save(file_path)

    except Exception as e:
        print(f"发生错误: {e}")


if __name__ == '__main__':
    # 示例字典数据
    data = {
        "candle_begin_time_GMT8": ["2022-01-01", "2021-01-02"],
        "open": [100, 105],
        "high": [110, 115],
        "low": [90, 95],
        "close": [105, 110],
        "long_shadow": [5, 5],
        "short_shadow": [2, 2],
        "body": [10, 10],
        "long": [12, 12]
    }
    # 调用方法将数据写入Excel
    # write_dict_to_existing_excel(data, "Pinbar")
    delete_last_row()
