import datetime
import os

import openpyxl


class MakeExcel:
    def __init__(self, sheet_name="Sheet1"):
        # 字典键与Excel列名的映射关系
        self.key_to_header = {
            "candle_begin_time_GMT8": "日期",
            "trading_target": "币种",
            "open": "开盘价",
            "high": "最高价",
            "low": "最低价",
            "close": "收盘价",
            "long_shadow": "长影线",
            "short_shadow": "短影线",
            "body": "实体长",
            "long": "总长"
        }
        # 获取当前时间并格式化
        now = datetime.datetime.now()
        current_time = now.strftime("%Y%m%d")
        # 文件路径"E:\PrivateFile\关于交易\统计\K线形态统计.xlsx"
        self.file_path = f"E:\\PrivateFile\\关于交易\\统计\\{current_time}.xlsx"
        self.sheet_name = sheet_name

    def write_dict_to_existing_excel(self, data_dict):
        """
        将字典数据写入到指定的现有Excel文件的工作表中。如果文件或工作表不存在，则创建它们。
        :param data_dict: 要写入Excel的字典数据，键为数据列名，值为要写入的数据列表
        :param sheet_name: 要写入数据的工作表名称，默认为 "Sheet1"
        """

        if not os.path.exists(self.file_path):
            # 如果文件不存在，则创建一个新的工作簿和工作表
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = self.sheet_name

            # 写入列名
            for col, column_name in enumerate(self.key_to_header.values(), start=1):
                sheet.cell(row=1, column=col, value=column_name)
        else:
            # 加载现有的工作簿
            wb = openpyxl.load_workbook(self.file_path)

            # 获取指定名称的工作表，如果不存在则创建
            if self.sheet_name in wb.sheetnames:
                sheet = wb[self.sheet_name]
            else:
                sheet = wb.create_sheet(self.sheet_name)
                # 写入列名
                for col, column_name in enumerate(self.key_to_header.values(), start=1):
                    sheet.cell(row=1, column=col, value=column_name)

        # 获取现有表头并创建一个列名到列号的映射
        header_to_col = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is not None:
                header_to_col[cell_value] = col

        # 将数据写入对应的列名下
        for key in data_dict.keys():
            if key in self.key_to_header.keys():
                column_name = self.key_to_header[key]
                if column_name in header_to_col:
                    col_num = header_to_col[column_name]
                    # 找到该列的第一个空单元格
                    row_num = 2
                    while sheet.cell(row=row_num, column=col_num).value is not None:
                        row_num += 1
                    # 写入数据
                    sheet.cell(row=row_num, column=col_num, value=data_dict[key])
        # 保存工作簿
        wb.save(self.file_path)
        # print(f"数据已成功写入到 {self.file_path} 的 {self.sheet_name} 工作表中")

    def delete_last_row(self):
        """
        删除Excel文件中指定工作表的最后一行内容。
        :param file_path: Excel文件的路径
        :param sheet_name: 要处理的工作表名称，默认为 "Sheet1"
        """
        if not os.path.exists(self.file_path):
            print(f"文件 {self.file_path} 不存在。")
            return

        try:
            # 加载现有的工作簿
            wb = openpyxl.load_workbook(self.file_path)

            # 检查工作表是否存在
            if self.sheet_name not in wb.sheetnames:
                print(f"工作表 {self.sheet_name} 不存在。")
                return

            sheet = wb[self.sheet_name]

            # 找到最后一行有数据的行号
            last_row = sheet.max_row
            while last_row > 1:
                if any(sheet.cell(row=last_row, column=col).value is not None for col in
                       range(1, sheet.max_column + 1)):
                    break
                last_row -= 1

            if last_row > 1:
                # 清除最后一行的数据
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=last_row, column=col).value = None
                print(f"已删除 {self.sheet_name} 工作表的最后一行内容。")
            else:
                print(f"{self.sheet_name} 工作表没有可删除的行。")

            # 保存工作簿
            wb.save(self.file_path)

        except Exception as e:
            print(f"发生错误: {e}")
